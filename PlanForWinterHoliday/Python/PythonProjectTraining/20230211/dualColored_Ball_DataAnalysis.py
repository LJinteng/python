import requests
import pprint
import time
import pandas as pd
import json
import openpyxl
from collections import Counter
#导入url以及headers
urls = "https://jc.zhcw.com/port/client_json.php?callback=jQuery1122010531339367835058_1676110388971&transactionType=10001001&lotteryId=1&issueCount=1000&startIssue=&endIssue=&startDate=&endDate=&type=0&pageNum=1&pageSize=1000&tt=0.5776440858451428&_=1676110388973"
headers = {
    'Cookie': 'Hm_lvt_692bd5f9c07d3ebd0063062fb0d7622f=1676109037; PHPSESSID=44n4p9u4v0q6pe2sl4l62dj412; Hm_lvt_12e4883fd1649d006e3ae22a39f97330=1676109037; _ga=GA1.2.1898445428.1676109038; _gid=GA1.2.713510709.1676109038; _gat_UA-66069030-3=1; Hm_lpvt_12e4883fd1649d006e3ae22a39f97330=1676110389; KLBRSID=13ce4968858adba085afff577d78760d|1676110388|1676110370; Hm_lpvt_692bd5f9c07d3ebd0063062fb0d7622f=1676110389',
    'Host': 'jc.zhcw.com',
    'Referer': 'https://www.zhcw.com/',
    'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/110.0.0.0 Safari/537.36'
}
#获取信息
response = requests.get(url = urls, headers = headers)
data = json.loads(response.text[43 : -1])["data"]
all_data = {}

for i in range(len(data)):
    ball_data = []
    date = data[i]['openTime']
    frontWinningNum = data[i]["frontWinningNum"].split()
    for j in range(len(frontWinningNum)):
        ball_data.append(int(frontWinningNum[j]))
    backWinningNum = data[i]["backWinningNum"]
    ball_data.append(int(backWinningNum))
    all_data[f'{date}'] = ball_data
    load_number = i / (len(data)+1)
    print('-----------正在获取数据中-%.2f%%-----------' % (load_number * 100))
    time.sleep(0.01)
print('-----------正在获取数据中-100.0%-----------')
#获取红球的列表
redball_list = []
for ball in all_data.values():
    redball_list.append(ball[0:6])


#获取蓝球的列表
blueball_list = []
for ball in all_data.values():
    blueball_list.append(ball[6])

redball_result = {}
for redball in redball_list:
    count = Counter(redball)
    for redball_k, redball_v in count.items():
        if redball_k in redball_result:
            redball_result[redball_k] += redball_v
        else:
            redball_result[redball_k] = redball_v


count = Counter(blueball_list)
blueball_result = {}
for blueball_k, blueball_v in count.items():
    if blueball_k in blueball_result:
        blueball_result[blueball_k] += blueball_v
    else:
        blueball_result[blueball_k] = blueball_v


workbook = openpyxl.Workbook()
sheet = workbook.active
sheet.title = 'get_dualcoloredball_sata'

# Write headers
sheet.cell(row=1, column=1, value='Redball_Value')
sheet.cell(row=1, column=2, value='Redball_Frequency')
sheet.cell(row=1, column=3, value='Blueball_Value')
sheet.cell(row=1, column=4, value='Blueball_Frequency')
#
# # Write data
row = 2
for k, v in redball_result.items():
    sheet.cell(row=row, column=1, value=k)
    sheet.cell(row=row, column=2, value=v)
    row += 1


row = 2
for k, v in blueball_result.items():
    sheet.cell(row=row, column=3, value=k)
    sheet.cell(row=row, column=4, value=v)
    row += 1
print('---------------正在计算数据中--------------')

time.sleep(1)
print('---------------正在计算数据中--------------')
# # Save the file
flimname =  'dualColored_ball.xlsx'
workbook.save(flimname)

df = pd.read_excel(flimname)
time.sleep(1)
print('---------------正在计算数据中--------------')
# 获取第一列数据
redball_value = df.iloc[:, 1].tolist()
redball_value = sorted(redball_value, reverse=True)
redball_value = redball_value[:6]
time.sleep(1)
blueball_value = max(df.iloc[:, 3].tolist())
print("----------------成功完成演算---------------")
redball = [key for key, value in redball_result.items() if value in redball_value]
blueball = [key for key, value in blueball_result.items() if value == blueball_value]
print(f'建议购买：红球{redball}\t蓝球{blueball}')